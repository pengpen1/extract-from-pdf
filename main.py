"""
简历信息提取工具
从PDF简历中自动提取姓名、电话、邮箱等关键信息，并导出为Excel文件
"""

from dataclasses import dataclass
from typing import List, Optional
from pathlib import Path


# ==================== 数据模型 ====================


@dataclass
class ResumeInfo:
    """简历信息数据类"""

    index: int  # 序号
    name: str  # 姓名
    gender: str  # 性别
    age: str  # 年龄
    date: str  # 时间
    phone: str  # 电话
    position: str  # 岗位
    location: str  # 地区
    salary: str  # 工资
    email: str  # 邮箱
    filename: str  # 文件名


@dataclass
class ProcessingResult:
    """处理结果统计"""

    total: int  # 总文件数
    success: int  # 成功处理数
    failed: int  # 失败数
    failed_files: List[str]  # 失败文件列表


# ==================== 自定义异常 ====================


class PDFExtractionError(Exception):
    """PDF提取异常"""

    pass


class ExcelExportError(Exception):
    """Excel导出异常"""

    pass


# ==================== 文件扫描模块 ====================


class FileScanner:
    """文件扫描器，负责递归扫描指定目录中的所有PDF文件"""

    def __init__(self, data_folder: str):
        """初始化文件扫描器

        Args:
            data_folder: 数据文件夹路径
        """
        self.data_folder = Path(data_folder)

        # 如果文件夹不存在，自动创建
        if not self.data_folder.exists():
            self.data_folder.mkdir(parents=True, exist_ok=True)
            print(f"已自动创建数据文件夹: {self.data_folder.absolute()}")

    def scan_pdf_files(self) -> List[Path]:
        """递归扫描所有PDF文件

        使用pathlib.Path.rglob递归查找所有.pdf文件

        Returns:
            PDF文件路径列表
        """
        # 使用rglob递归查找所有PDF文件（不区分大小写）
        pdf_files = []

        # 查找.pdf文件
        pdf_files.extend(self.data_folder.rglob("*.pdf"))

        # 查找.PDF文件（大写扩展名）
        pdf_files.extend(self.data_folder.rglob("*.PDF"))

        # 去重并排序
        pdf_files = sorted(set(pdf_files))

        return pdf_files


# ==================== PDF文本提取模块 ====================


class PDFExtractor:
    """PDF文本提取器，负责从PDF文件中提取文本内容"""

    def extract_text(self, pdf_path: Path) -> str:
        """提取PDF文本内容

        优先使用pdfplumber提取（更准确），失败时回退到PyPDF2。
        只提取前3页内容以提高处理速度。

        Args:
            pdf_path: PDF文件路径

        Returns:
            提取的文本内容

        Raises:
            PDFExtractionError: 当所有提取方法都失败时抛出
        """
        text = ""

        # 方法1: 尝试使用pdfplumber提取（主要方法）
        try:
            text = self._extract_with_pdfplumber(pdf_path)
            if text.strip():  # 如果成功提取到非空文本
                return text
        except Exception as e:
            # pdfplumber失败，记录错误但继续尝试备用方法
            print(f"  pdfplumber提取失败: {str(e)}")

        # 方法2: 回退到PyPDF2（备用方法）
        try:
            text = self._extract_with_pypdf2(pdf_path)
            if text.strip():  # 如果成功提取到非空文本
                return text
        except Exception as e:
            # PyPDF2也失败，抛出异常
            raise PDFExtractionError(f"无法提取PDF文本: {str(e)}")

        # 如果两种方法都没有提取到有效文本
        if not text.strip():
            raise PDFExtractionError("PDF文件为空或无法提取文本内容")

        return text

    def _extract_with_pdfplumber(self, pdf_path: Path) -> str:
        """使用pdfplumber提取PDF文本

        Args:
            pdf_path: PDF文件路径

        Returns:
            提取的文本内容
        """
        import pdfplumber

        text_parts = []

        try:
            with pdfplumber.open(pdf_path) as pdf:
                # 只提取前3页
                max_pages = min(3, len(pdf.pages))

                for i in range(max_pages):
                    page = pdf.pages[i]
                    page_text = page.extract_text()

                    if page_text:
                        text_parts.append(page_text)
        except Exception as e:
            raise PDFExtractionError(f"pdfplumber提取失败: {str(e)}")

        return "\n".join(text_parts)

    def _extract_with_pypdf2(self, pdf_path: Path) -> str:
        """使用PyPDF2提取PDF文本（备用方法）

        Args:
            pdf_path: PDF文件路径

        Returns:
            提取的文本内容
        """
        import PyPDF2

        text_parts = []

        try:
            with open(pdf_path, "rb") as file:
                pdf_reader = PyPDF2.PdfReader(file)

                # 只提取前3页
                max_pages = min(3, len(pdf_reader.pages))

                for i in range(max_pages):
                    page = pdf_reader.pages[i]
                    page_text = page.extract_text()

                    if page_text:
                        text_parts.append(page_text)
        except Exception as e:
            raise PDFExtractionError(f"PyPDF2提取失败: {str(e)}")

        return "\n".join(text_parts)


# ==================== 信息提取模块 ====================


class InfoExtractor:
    """信息提取器，负责从文本中识别和提取姓名、电话、邮箱"""

    def __init__(self):
        """初始化信息提取器"""
        import re

        self.re = re

        # 常见标题词，用于过滤非姓名文本
        self.common_title_words = {
            "个人简历",
            "求职简历",
            "简历",
            "个人信息",
            "基本信息",
            "求职意向",
            "工作经历",
            "教育经历",
            "项目经验",
            "自我评价",
            "技能特长",
            "联系方式",
            "应聘岗位",
            "期望职位",
            "个人资料",
            "前端工程师",
            "后端工程师",
            "前端开发",
            "后端开发",
            "全栈工程师",
            "开发工程师",
        }

    def parse_filename(self, filename: str) -> dict:
        """从文件名中解析信息

        文件名格式示例：【前端开发工程师_成都 8-12K】李志华 5年.pdf

        Args:
            filename: 文件名

        Returns:
            包含解析信息的字典
        """
        info = {
            "name": None,
            "position": None,
            "location": None,
            "salary": None,
        }

        # 移除.pdf扩展名和常见后缀
        name_without_ext = filename.replace(".pdf", "").replace(".PDF", "")
        name_without_ext = self.re.sub(r"的简历$", "", name_without_ext)  # 移除"的简历"

        # 模式1: 【岗位_地区 薪资】姓名 年限
        pattern1 = r"【([^_]+)_([^\s]+)\s+([^】]+)】\s*([^\s]+)"
        match = self.re.search(pattern1, name_without_ext)
        if match:
            info["position"] = match.group(1).strip()
            info["location"] = match.group(2).strip()
            info["salary"] = match.group(3).strip()
            info["name"] = match.group(4).strip()
            return info

        # 模式2: 姓名在文件名中（简单模式）
        # 尝试提取2-4个连续中文字符作为姓名
        chinese_pattern = r"[\u4e00-\u9fff]{2,4}"
        matches = self.re.findall(chinese_pattern, name_without_ext)
        if matches:
            # 取第一个有效的姓名
            for match in matches:
                if self._is_valid_name(match):
                    info["name"] = match
                    break

        return info

    def extract_phone(self, text: str) -> Optional[str]:
        """提取手机号码

        支持多种格式：
        - 标准格式：13812345678
        - 带括号：(+86) 138-1234-5678
        - 带横线：138-1234-5678
        - 带空格：138 1234 5678

        Args:
            text: 简历文本

        Returns:
            手机号码字符串，未找到时返回None
        """
        if not text:
            return None

        # 先尝试提取所有可能包含手机号的文本段
        # 查找包含1开头数字的所有片段
        potential_phones = self.re.findall(r"[\(\+\d\s\-]{11,20}", text)

        for segment in potential_phones:
            # 提取纯数字
            digits = self.re.sub(r"[^\d]", "", segment)
            # 验证是否为11位手机号
            if len(digits) == 11 and digits[0] == "1" and digits[1] in "3456789":
                return digits

        # 如果上面没找到，使用标准模式
        pattern = r"\b1[3-9]\d{9}\b"
        match = self.re.search(pattern, text)
        if match:
            return match.group(0)

        return None

    def extract_email(self, text: str) -> Optional[str]:
        """提取邮箱地址

        使用正则表达式匹配标准邮箱地址格式。
        当存在多个邮箱时，返回第一个匹配的结果。

        Args:
            text: 简历文本

        Returns:
            邮箱地址字符串，未找到时返回None
        """
        if not text:
            return None

        # 正则表达式: 匹配标准邮箱格式
        # 用户名部分: 字母、数字、点、下划线、百分号、加号、减号（至少3个字符）
        # @符号
        # 域名部分: 字母、数字、点、减号
        # 顶级域名: 至少2个字母
        pattern = r"[a-zA-Z0-9][a-zA-Z0-9._%+-]{2,}@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"

        # 查找所有匹配
        matches = self.re.findall(pattern, text)

        # 返回最长的邮箱（避免截断问题）
        if matches:
            return max(matches, key=len)

        return None

    def extract_name(self, text: str) -> Optional[str]:
        """提取姓名

        使用多策略组合方法提取姓名：
        1. 关键词模式匹配（优先）
        2. 位置启发式策略（回退）

        Args:
            text: 简历文本

        Returns:
            姓名字符串，未找到时返回None
        """
        if not text:
            return None

        # 策略1: 关键词模式匹配
        name = self._extract_name_by_keyword(text)
        if name:
            return name

        # 策略2: 位置启发式策略
        name = self._extract_name_by_position(text)
        if name:
            return name

        return None

    def _extract_name_by_keyword(self, text: str) -> Optional[str]:
        """通过关键词模式提取姓名

        搜索"姓名："、"姓 名："、"Name:"等关键词后的内容

        Args:
            text: 简历文本

        Returns:
            提取的姓名，未找到时返回None
        """
        # 定义关键词模式列表
        keyword_patterns = [
            r"姓\s*名\s*[：:]\s*([^\s\n]{2,4})",  # 姓名：、姓 名：
            r"姓\s*名\s*[：:]\s*([^\s\n]{2,4})",  # 姓名:（英文冒号）
            r"Name\s*[：:]\s*([^\s\n]{2,4})",  # Name:、Name：
            r"name\s*[：:]\s*([^\s\n]{2,4})",  # name:、name：
            r"名\s*字\s*[：:]\s*([^\s\n]{2,4})",  # 名字：
        ]

        for pattern in keyword_patterns:
            match = self.re.search(pattern, text, self.re.IGNORECASE)
            if match:
                candidate = match.group(1).strip()
                # 验证候选姓名
                if self._is_valid_name(candidate):
                    return candidate

        return None

    def _extract_name_by_position(self, text: str) -> Optional[str]:
        """通过位置启发式策略提取姓名

        在简历前100字符中查找2-4个连续中文字符作为候选姓名
        优先查找第一行或前几行的独立中文词

        Args:
            text: 简历文本

        Returns:
            提取的姓名，未找到时返回None
        """
        # 只在前100字符中查找（缩小范围，提高准确性）
        search_text = text[:100]

        # 按行分割
        lines = search_text.split("\n")

        # 优先检查前3行
        for line in lines[:3]:
            line = line.strip()
            # 查找2-4个连续中文字符
            pattern = r"^[\u4e00-\u9fff]{2,4}$"
            match = self.re.match(pattern, line)
            if match:
                candidate = match.group(0)
                if self._is_valid_name(candidate):
                    return candidate

        # 如果前3行没找到，在整个搜索文本中查找
        pattern = r"[\u4e00-\u9fff]{2,4}"
        matches = self.re.findall(pattern, search_text)

        # 遍历所有匹配，找到第一个有效的姓名
        for candidate in matches:
            if self._is_valid_name(candidate):
                return candidate

        return None

    def _is_valid_name(self, candidate: str) -> bool:
        """验证候选文本是否为有效姓名

        过滤规则：
        - 排除常见标题词
        - 排除包含数字的文本
        - 排除包含特殊符号的文本
        - 长度必须在2-4个字符之间

        Args:
            candidate: 候选姓名文本

        Returns:
            True表示有效姓名，False表示无效
        """
        if not candidate:
            return False

        # 去除首尾空白
        candidate = candidate.strip()

        # 检查长度（中文姓名通常2-4个字）
        if len(candidate) < 2 or len(candidate) > 4:
            return False

        # 排除常见标题词
        if candidate in self.common_title_words:
            return False

        # 排除包含数字的文本
        if self.re.search(r"\d", candidate):
            return False

        # 排除包含特殊符号的文本（允许中文字符）
        # 只允许纯中文字符
        if not self.re.match(r"^[\u4e00-\u9fff]+$", candidate):
            return False

        return True

    def extract_gender(self, text: str) -> Optional[str]:
        """提取性别

        搜索"性别："、"性 别："等关键词后的内容
        也尝试从其他线索推断（如"先生"、"女士"等）

        Args:
            text: 简历文本

        Returns:
            性别字符串（男/女），未找到时返回None
        """
        if not text:
            return None

        # 定义关键词模式列表
        patterns = [
            r"性\s*别\s*[：:]\s*(男|女)",
            r"Gender\s*[：:]\s*(男|女|Male|Female|male|female)",
            r"[：:]\s*(男|女)\s*[|/\n]",  # 匹配 "：男 |" 这种格式
        ]

        for pattern in patterns:
            match = self.re.search(pattern, text, self.re.IGNORECASE)
            if match:
                gender = match.group(1).strip()
                # 统一转换为中文
                if gender.lower() in ["male", "男"]:
                    return "男"
                elif gender.lower() in ["female", "女"]:
                    return "女"

        # 尝试从称呼推断（在前200字符中）
        search_text = text[:200]
        if "先生" in search_text or "Mr" in search_text:
            return "男"
        elif "女士" in search_text or "Ms" in search_text or "Miss" in search_text:
            return "女"

        return None

    def extract_age(self, text: str) -> Optional[str]:
        """提取年龄

        搜索"年龄："、"Age:"等关键词后的数字，或者"XX岁"格式

        Args:
            text: 简历文本

        Returns:
            年龄字符串，未找到时返回None
        """
        if not text:
            return None

        # 定义关键词模式列表
        patterns = [
            r"年\s*龄\s*[：:]\s*(\d{1,2})",
            r"Age\s*[：:]\s*(\d{1,2})",
            r"(\d{2})岁",  # 匹配"26岁"这种格式
        ]

        for pattern in patterns:
            match = self.re.search(pattern, text, self.re.IGNORECASE)
            if match:
                age = match.group(1).strip()
                # 验证年龄范围合理性（18-70岁）
                if age.isdigit() and 18 <= int(age) <= 70:
                    return age

        return None

    def extract_date(self, text: str) -> Optional[str]:
        """提取日期/时间

        搜索简历中的日期信息（如更新时间、出生日期等）

        Args:
            text: 简历文本

        Returns:
            日期字符串，未找到时返回None
        """
        if not text:
            return None

        # 定义日期模式列表
        patterns = [
            r"更新时间\s*[：:]\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}日?)",
            r"出生日期\s*[：:]\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}日?)",
            r"(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}日?)",
        ]

        for pattern in patterns:
            match = self.re.search(pattern, text)
            if match:
                return match.group(1).strip()

        return None

    def extract_position(self, text: str) -> Optional[str]:
        """提取应聘岗位

        搜索"应聘岗位："、"期望职位："等关键词后的内容
        如果内容包含城市名，只提取岗位部分

        Args:
            text: 简历文本

        Returns:
            岗位字符串，未找到时返回None
        """
        if not text:
            return None

        # 定义关键词模式列表
        patterns = [
            r"应聘岗位\s*[：:]\s*([^\n]+)",
            r"期望职位\s*[：:]\s*([^\n]+)",
            r"求职意向\s*[：:]\s*([^\n]+)",
            r"目标职位\s*[：:]\s*([^\n]+)",
            r"Position\s*[：:]\s*([^\n]+)",
        ]

        for pattern in patterns:
            match = self.re.search(pattern, text, self.re.IGNORECASE)
            if match:
                full_text = match.group(1).strip()

                # 如果包含常见分隔符（空格、|、/等），尝试分离岗位和地区
                # 例如："Java 成都" 或 "Java | 成都"
                separators = [r"\s+", r"\|", r"/", r"·"]

                for sep in separators:
                    parts = self.re.split(sep, full_text)
                    if len(parts) >= 2:
                        # 取第一部分作为岗位
                        position = parts[0].strip()
                        # 验证岗位不是城市名
                        if position and not self._is_city_name(position):
                            return position

                # 如果没有分隔符，返回整个文本（但要验证长度合理）
                if len(full_text) <= 15:
                    return full_text

        return None

    def extract_location(self, text: str) -> Optional[str]:
        """提取地区/城市

        搜索"工作地点："、"期望城市："等关键词后的内容
        或从"求职意向"中提取城市部分
        只返回识别的城市名，避免返回过多无关内容

        Args:
            text: 简历文本

        Returns:
            地区字符串，未找到时返回None
        """
        if not text:
            return None

        # 定义关键词模式列表
        patterns = [
            r"期望城市\s*[：:]\s*([^\n|]+)",
            r"工作地点\s*[：:]\s*([^\n|]+)",
            r"期望地点\s*[：:]\s*([^\n|]+)",
        ]

        for pattern in patterns:
            match = self.re.search(pattern, text, self.re.IGNORECASE)
            if match:
                location_text = match.group(1).strip()
                # 从文本中提取城市名
                city = self._extract_city_from_text(location_text)
                if city:
                    return city

        # 如果没有找到，尝试从"求职意向"中提取城市
        intention_pattern = r"求职意向\s*[：:]\s*([^\n]+)"
        match = self.re.search(intention_pattern, text)
        if match:
            full_text = match.group(1).strip()

            # 尝试分离岗位和地区
            separators = [r"\s+", r"\|", r"/", r"·"]

            for sep in separators:
                parts = self.re.split(sep, full_text)
                if len(parts) >= 2:
                    # 取第二部分作为地区
                    location = parts[1].strip()
                    # 验证是否为城市名
                    if location and self._is_city_name(location):
                        return location

        return None

    def _extract_city_from_text(self, text: str) -> Optional[str]:
        """从文本中提取城市名

        Args:
            text: 包含城市信息的文本

        Returns:
            城市名，未找到时返回None
        """
        # 常见城市名列表
        cities = {
            "北京",
            "上海",
            "广州",
            "深圳",
            "成都",
            "重庆",
            "杭州",
            "武汉",
            "西安",
            "天津",
            "南京",
            "苏州",
            "长沙",
            "郑州",
            "沈阳",
            "青岛",
            "宁波",
            "东莞",
            "无锡",
            "佛山",
            "合肥",
            "昆明",
            "福州",
            "厦门",
            "哈尔滨",
            "济南",
            "温州",
            "长春",
            "石家庄",
            "常州",
            "泉州",
            "南宁",
            "贵阳",
            "南昌",
            "南通",
            "金华",
            "徐州",
            "太原",
            "嘉兴",
            "烟台",
            "惠州",
            "保定",
            "台州",
            "中山",
            "绍兴",
            "乌鲁木齐",
            "潍坊",
            "兰州",
        }

        # 在文本中查找城市名
        for city in cities:
            if city in text:
                return city

        return None

    def _is_city_name(self, text: str) -> bool:
        """判断文本是否为城市名

        Args:
            text: 待判断的文本

        Returns:
            True表示是城市名，False表示不是
        """
        # 常见城市名列表（可以根据需要扩展）
        cities = {
            "北京",
            "上海",
            "广州",
            "深圳",
            "成都",
            "重庆",
            "杭州",
            "武汉",
            "西安",
            "天津",
            "南京",
            "苏州",
            "长沙",
            "郑州",
            "沈阳",
            "青岛",
            "宁波",
            "东莞",
            "无锡",
            "佛山",
            "合肥",
            "昆明",
            "福州",
            "厦门",
            "哈尔滨",
            "济南",
            "温州",
            "长春",
            "石家庄",
            "常州",
            "泉州",
            "南宁",
            "贵阳",
            "南昌",
            "南通",
            "金华",
            "徐州",
            "太原",
            "嘉兴",
            "烟台",
            "惠州",
            "保定",
            "台州",
            "中山",
            "绍兴",
            "乌鲁木齐",
            "潍坊",
            "兰州",
        }

        return text in cities

    def extract_salary(self, text: str) -> Optional[str]:
        """提取期望薪资

        搜索"期望薪资："、"薪资要求："等关键词后的内容
        只提取薪资数字部分，避免包含其他无关信息

        Args:
            text: 简历文本

        Returns:
            薪资字符串，未找到时返回None
        """
        if not text:
            return None

        # 定义关键词模式列表
        patterns = [
            r"期望薪资\s*[：:]\s*([^\n|]+)",
            r"薪资要求\s*[：:]\s*([^\n|]+)",
            r"期望工资\s*[：:]\s*([^\n|]+)",
        ]

        for pattern in patterns:
            match = self.re.search(pattern, text, self.re.IGNORECASE)
            if match:
                salary_text = match.group(1).strip()
                # 提取薪资格式：数字-数字K 或 数字k~数字k
                salary = self._extract_salary_from_text(salary_text)
                if salary:
                    return salary

        return None

    def _extract_salary_from_text(self, text: str) -> Optional[str]:
        """从文本中提取薪资数字

        支持格式：
        - 8-12K
        - 8k~12k
        - 8000-12000

        Args:
            text: 包含薪资信息的文本

        Returns:
            薪资字符串，未找到时返回None
        """
        # 匹配薪资格式
        patterns = [
            r"(\d+[-~]\d+[kK])",  # 8-12K 或 8~12k
            r"(\d+[kK][-~]\d+[kK])",  # 8k-12k
            r"(\d+[-~]\d+)",  # 8000-12000
        ]

        for pattern in patterns:
            match = self.re.search(pattern, text)
            if match:
                salary = match.group(1)
                # 统一格式，转换为大写K
                salary = salary.replace("k", "K")
                return salary

        return None


# ==================== Excel导出模块 ====================


class ExcelExporter:
    """Excel导出器，负责将提取结果导出为Excel文件"""

    def __init__(self, output_dir: Path = None):
        """初始化Excel导出器

        Args:
            output_dir: 输出目录路径，默认为当前目录
        """
        if output_dir is None:
            # 默认输出到当前目录
            self.output_dir = Path.cwd()
        else:
            self.output_dir = Path(output_dir)

        # 确保输出目录存在
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export(self, data: List[ResumeInfo]) -> Path:
        """导出数据到Excel文件

        Args:
            data: 简历信息列表

        Returns:
            生成的Excel文件路径

        Raises:
            ExcelExportError: 导出失败时抛出
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            from datetime import datetime

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "简历信息"

            # 定义表头
            headers = [
                "序号",
                "姓名",
                "性别",
                "年龄",
                "时间",
                "电话",
                "岗位",
                "地区",
                "工资",
                "邮箱",
                "文件名",
            ]

            # 写入表头
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                # 设置表头样式：粗体、居中
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 写入数据行
            for row_idx, resume_info in enumerate(data, start=2):
                ws.cell(row=row_idx, column=1, value=resume_info.index)
                ws.cell(row=row_idx, column=2, value=resume_info.name)
                ws.cell(row=row_idx, column=3, value=resume_info.gender)
                ws.cell(row=row_idx, column=4, value=resume_info.age)
                ws.cell(row=row_idx, column=5, value=resume_info.date)
                ws.cell(row=row_idx, column=6, value=resume_info.phone)
                ws.cell(row=row_idx, column=7, value=resume_info.position)
                ws.cell(row=row_idx, column=8, value=resume_info.location)
                ws.cell(row=row_idx, column=9, value=resume_info.salary)
                ws.cell(row=row_idx, column=10, value=resume_info.email)
                ws.cell(row=row_idx, column=11, value=resume_info.filename)

            # 自动调整列宽
            self._adjust_column_width(ws)

            # 生成时间戳文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"简历提取结果_{timestamp}.xlsx"
            output_path = self.output_dir / filename

            # 保存Excel文件
            wb.save(output_path)

            return output_path

        except ImportError as e:
            raise ExcelExportError(f"缺少openpyxl库: {str(e)}")
        except Exception as e:
            raise ExcelExportError(f"Excel导出失败: {str(e)}")

    def _adjust_column_width(self, worksheet):
        """自动调整列宽

        根据单元格内容自动调整每列的宽度，使内容完整显示

        Args:
            worksheet: openpyxl工作表对象
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # 获取列字母（A, B, C...）

            for cell in column:
                try:
                    # 计算单元格内容长度
                    if cell.value:
                        # 中文字符按2个字符计算，英文按1个字符计算
                        cell_value = str(cell.value)
                        # 简单估算：中文字符占用更多空间
                        length = sum(
                            2 if "\u4e00" <= char <= "\u9fff" else 1
                            for char in cell_value
                        )
                        max_length = max(max_length, length)
                except:
                    pass

            # 设置列宽（加上一些边距）
            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
            worksheet.column_dimensions[column_letter].width = adjusted_width


# ==================== 主控制器 ====================


class ResumeExtractorApp:
    """简历信息提取应用主控制器，负责协调各组件并控制整体流程"""

    def __init__(self, data_folder: str = "datas"):
        """初始化应用

        Args:
            data_folder: 数据文件夹名称，默认为"datas"
        """
        self.data_folder = data_folder
        self.file_scanner = FileScanner(data_folder)
        self.pdf_extractor = PDFExtractor()
        self.info_extractor = InfoExtractor()
        self.excel_exporter = ExcelExporter()

    def run(self):
        """运行主流程

        流程步骤：
        1. 显示启动欢迎界面
        2. 扫描PDF文件
        3. 检查是否有文件需要处理
        4. 逐个处理文件（显示进度）
        5. 收集提取结果
        6. 导出Excel
        7. 显示统计信息
        8. 等待用户按键退出
        """
        # 步骤1: 显示启动欢迎界面
        self._show_welcome()

        # 步骤2: 扫描PDF文件
        print("\n正在扫描文件...")
        pdf_files = self.file_scanner.scan_pdf_files()

        # 步骤3: 检查是否有文件需要处理
        if len(pdf_files) == 0:
            print(f"\n未在 '{self.data_folder}' 文件夹中找到PDF文件。")
            print(f"请将PDF简历文件放入 '{self.data_folder}' 文件夹后重新运行程序。")
            self._wait_for_exit()
            return

        print(f"找到 {len(pdf_files)} 个PDF文件\n")

        # 步骤4-5: 处理文件并收集结果
        print("开始处理...\n")
        resume_list = []
        failed_files = []

        for index, pdf_file in enumerate(pdf_files, start=1):
            # 显示进度
            print(f"[{index}/{len(pdf_files)}] 正在处理: {pdf_file.name} ... ", end="")

            try:
                # 提取PDF文本
                text = self.pdf_extractor.extract_text(pdf_file)

                # 从文件名中解析信息（作为补充）
                filename_info = self.info_extractor.parse_filename(pdf_file.name)

                # 提取信息（优先使用PDF内容，文件名作为补充）
                name = (
                    self.info_extractor.extract_name(text)
                    or filename_info.get("name")
                    or ""
                )
                gender = self.info_extractor.extract_gender(text) or ""
                age = self.info_extractor.extract_age(text) or ""
                date = self.info_extractor.extract_date(text) or ""
                phone = self.info_extractor.extract_phone(text) or ""
                position = (
                    self.info_extractor.extract_position(text)
                    or filename_info.get("position")
                    or ""
                )
                location = (
                    self.info_extractor.extract_location(text)
                    or filename_info.get("location")
                    or ""
                )
                salary = (
                    self.info_extractor.extract_salary(text)
                    or filename_info.get("salary")
                    or ""
                )
                email = self.info_extractor.extract_email(text) or ""

                # 创建简历信息对象
                resume_info = ResumeInfo(
                    index=index,
                    name=name,
                    gender=gender,
                    age=age,
                    date=date,
                    phone=phone,
                    position=position,
                    location=location,
                    salary=salary,
                    email=email,
                    filename=pdf_file.name,
                )

                resume_list.append(resume_info)
                print("✓")

            except PDFExtractionError as e:
                # 记录失败文件
                failed_files.append((pdf_file.name, str(e)))
                print("✗")
            except Exception as e:
                # 捕获其他未预期的错误
                failed_files.append((pdf_file.name, f"未知错误: {str(e)}"))
                print("✗")

        # 步骤6: 导出Excel
        if resume_list:
            print("\n正在导出Excel文件...")
            try:
                output_path = self.excel_exporter.export(resume_list)
                print(f"✓ 导出成功")
            except ExcelExportError as e:
                print(f"✗ 导出失败: {e}")
                output_path = None
        else:
            print("\n没有成功提取的数据，跳过Excel导出。")
            output_path = None

        # 步骤7: 显示统计信息
        self._show_statistics(
            total=len(pdf_files),
            success=len(resume_list),
            failed=len(failed_files),
            failed_files=failed_files,
            output_path=output_path,
        )

        # 步骤8: 等待用户按键退出
        self._wait_for_exit()

    def _show_welcome(self):
        """显示启动欢迎界面"""
        print("=" * 60)
        print("    简历信息提取工具 v1.0")
        print("=" * 60)
        print("使用说明：")
        print("1. 将PDF简历放入 'datas' 文件夹")
        print("2. 双击运行本程序")
        print("3. 等待处理完成，查看生成的Excel文件")
        print("=" * 60)

    def _show_statistics(
        self,
        total: int,
        success: int,
        failed: int,
        failed_files: List[tuple],
        output_path: Path = None,
    ):
        """显示处理完成后的统计信息

        Args:
            total: 总文件数
            success: 成功处理数
            failed: 失败数
            failed_files: 失败文件列表，每项为(文件名, 错误原因)元组
            output_path: 输出Excel文件路径
        """
        print("\n" + "=" * 60)
        print("处理完成！")
        print("=" * 60)
        print(f"总文件数: {total}")
        print(f"成功: {success}")
        print(f"失败: {failed}")

        # 显示失败文件列表
        if failed_files:
            print("\n失败文件:")
            for filename, error in failed_files:
                print(f"  - {filename} ({error})")

        # 显示输出文件路径
        if output_path:
            print(f"\n输出文件: {output_path.absolute()}")

        print("=" * 60)

    def _wait_for_exit(self):
        """等待用户按键后退出"""
        input("\n按回车键退出...")


# ==================== 主程序入口 ====================

if __name__ == "__main__":
    try:
        # 创建应用实例并运行
        app = ResumeExtractorApp()
        app.run()
    except KeyboardInterrupt:
        # 用户中断程序（Ctrl+C）
        print("\n\n程序已被用户中断。")
        input("\n按回车键退出...")
    except Exception as e:
        # 捕获所有未处理的致命错误
        print("\n" + "=" * 60)
        print("程序运行出错！")
        print("=" * 60)
        print(f"错误类型: {type(e).__name__}")
        print(f"错误信息: {str(e)}")
        print("\n请检查以下可能的原因：")
        print("1. 确保 'datas' 文件夹存在且有读取权限")
        print("2. 确保有足够的磁盘空间")
        print("3. 确保PDF文件未被其他程序占用")
        print("4. 如果问题持续，请联系技术支持")
        print("=" * 60)
        input("\n按回车键退出...")
